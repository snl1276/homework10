from openpyxl import load_workbook
import csv


wb_val = load_workbook(filename='tabl.xlsx', data_only=True)
pointSheets = wb_val.sheetnames
sheet_val = wb_val[pointSheets[0]]


with open('output.csv', 'w', newline='') as csvfile:
    writer = csv.writer(
        csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for element in sheet_val:
        res = []
        for el in element:
            try:
                res.append(el.value)
            except ValueError:
                res.append(el.value)
            except TypeError:
                res.append('')
        writer.writerow(res)
